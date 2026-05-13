import os

from chunking.strategy import BaseChunkingStrategy, Chunk
from config import settings


class ExcelChunkingStrategy(BaseChunkingStrategy):
    def __init__(
        self,
        window_rows: int | None = None,
        window_overlap: int | None = None,
        max_chunk_size: int | None = None,
    ):
        self._window_rows = window_rows or getattr(settings, "excel_window_rows", 20)
        self._window_overlap = window_overlap or getattr(settings, "excel_window_overlap", 5)
        self._max_chunk_size = max_chunk_size or settings.chunk_size

    def chunk(self, text: str, metadata: dict | None = None) -> list[Chunk]:
        meta = dict(metadata or {})
        file_path = meta.get("source", "")
        if not file_path or not os.path.exists(file_path):
            return []

        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        chunks: list[Chunk] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]
            if not data_rows:
                continue

            pos = 0
            while pos < len(data_rows):
                window_end = min(pos + self._window_rows, len(data_rows))
                window = data_rows[pos:window_end]

                lines = [
                    f"Sheet: {sheet_name}",
                    f"Headers: {', '.join(headers)}",
                ]
                for r_idx, row in enumerate(window, start=pos + 2):
                    vals = [str(v) if v is not None else "" for v in row]
                    lines.append(f"Row {r_idx}: {' | '.join(vals)}")

                block = "\n".join(lines)

                if len(block) <= self._max_chunk_size:
                    chunk_meta = {
                        **meta,
                        "sheet_name": sheet_name,
                        "row_start": str(pos + 2),
                        "row_end": str(window_end + 1),
                    }
                    chunks.append(Chunk(text=block, metadata=chunk_meta))
                else:
                    half = self._window_rows // 2
                    sub_pos = pos
                    while sub_pos < window_end:
                        sub_end = min(sub_pos + max(half, 1), window_end)
                        sub_window = data_rows[sub_pos:sub_end]
                        sub_lines = [
                            f"Sheet: {sheet_name}",
                            f"Headers: {', '.join(headers)}",
                        ]
                        for r_idx, row in enumerate(sub_window, start=sub_pos + 2):
                            vals = [str(v) if v is not None else "" for v in row]
                            sub_lines.append(f"Row {r_idx}: {' | '.join(vals)}")
                        sub_block = "\n".join(sub_lines)

                        chunk_meta = {
                            **meta,
                            "sheet_name": sheet_name,
                            "row_start": str(sub_pos + 2),
                            "row_end": str(sub_end + 1),
                        }
                        chunks.append(Chunk(text=sub_block, metadata=chunk_meta))
                        sub_pos = sub_end

                next_pos = pos + self._window_rows - self._window_overlap
                if next_pos <= pos:
                    next_pos = pos + 1
                pos = next_pos

        wb.close()
        return chunks
